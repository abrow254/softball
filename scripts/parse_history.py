#!/usr/bin/env python3
"""Parse the team's Excel workbook into normalized season totals.

The workbook holds SEASON AGGREGATES (one row per player per season), with a
Regulars/Ringers split in newer seasons. This script reads every season sheet,
extracts the raw counting stats, dedups players across seasons, validates the
per-season sums against the All-Time Stats sheet, and emits a JSON blob the SQL
generator consumes.
"""
import json
import sys
import uuid
from openpyxl import load_workbook

PATH = '/Users/aaronbrown/Library/Mobile Documents/com~apple~CloudDocs/Downloads/The Softball Team - Stats.xlsx'

# (sheet name, year, term, is_current)
SEASON_SHEETS = [
    ('2023 Summer', 2023, 'Summer', False),
    ('2023 Fall',   2023, 'Fall',   False),
    ('2024 Summer', 2024, 'Summer', False),
    ('2024 Fall',   2024, 'Fall',   False),
    ('2025 Summer', 2025, 'Summer', False),
    ('2026 Summer', 2026, 'Summer', True),
]

# Header label -> our field. Only raw counting stats; rates are derived in SQL.
COUNT_COLS = {
    '1B': 'singles', '2B': 'doubles', '3B': 'triples', 'HR': 'hr',
    'AB': 'ab', 'FC': 'fc', 'BB': 'bb', 'HBP': 'hbp', 'ROE': 'roe',
    'RBI': 'rbi', 'Runs': 'runs', 'K': 'k', 'GP': 'gp',
}

NS = uuid.UUID('00000000-0000-0000-0000-00000000abcd')  # stable namespace


def player_id(name): return str(uuid.uuid5(NS, 'player:' + name))
def season_id(label): return str(uuid.uuid5(NS, 'season:' + label))
def game_id(label):   return str(uuid.uuid5(NS, 'game:' + label))


def num(v):
    if v is None:
        return 0
    if isinstance(v, str):
        return 0  # '#REF!', '#VALUE!', etc.
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return 0


def parse_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    # map header label -> column index
    idx = {}
    for i, h in enumerate(header):
        if isinstance(h, str) and h.strip() in COUNT_COLS:
            idx[COUNT_COLS[h.strip()]] = i

    players = []
    is_ringer = False
    for r in rows[1:]:
        c0 = r[0]
        if c0 is None:
            continue
        if not isinstance(c0, str):
            continue
        name = c0.strip()
        if name == '':
            continue
        if name == 'Ringer':
            is_ringer = True
            continue
        if name in ('Totals', 'Average') or name.startswith('*'):
            continue
        rec = {'name': name, 'is_regular': not is_ringer}
        for field in COUNT_COLS.values():
            rec[field] = num(r[idx[field]]) if field in idx else 0
        # skip ghost rows with no offensive activity at all
        if rec['ab'] == 0 and rec['singles'] == 0 and rec['doubles'] == 0 \
           and rec['triples'] == 0 and rec['hr'] == 0 and rec['k'] == 0:
            continue
        players.append(rec)
    return players


def main():
    wb = load_workbook(PATH, data_only=True, read_only=True)

    seasons = []
    all_players = {}      # name -> {is_regular (latest), ...}
    career_sum = {}       # name -> {field: total}

    for sheet, year, term, is_current in SEASON_SHEETS:
        label = f'{year} {term}'
        recs = parse_sheet(wb[sheet])
        seasons.append({
            'label': label, 'year': year, 'term': term,
            'is_current': is_current,
            'season_id': season_id(label), 'game_id': game_id(label),
            'players': recs,
        })
        for rec in recs:
            n = rec['name']
            # latest-season status wins (sheets are chronological)
            all_players[n] = rec['is_regular']
            cs = career_sum.setdefault(n, {f: 0 for f in COUNT_COLS.values()})
            for f in COUNT_COLS.values():
                cs[f] += rec[f]

    players = [
        {'name': n, 'is_regular': all_players[n], 'player_id': player_id(n)}
        for n in sorted(all_players)
    ]

    # ---- validate against All-Time Stats sheet ----
    at = wb['All-Time Stats']
    at_rows = list(at.iter_rows(values_only=True))
    at_header = list(at_rows[0])
    at_idx = {h.strip(): i for i, h in enumerate(at_header) if isinstance(h, str)}
    mismatches = []
    for r in at_rows[1:]:
        if not isinstance(r[0], str):
            continue
        n = r[0].strip()
        if n not in career_sum:
            mismatches.append(f'{n}: in All-Time but not parsed from seasons')
            continue
        for label, field in [('1B', 'singles'), ('2B', 'doubles'),
                              ('3B', 'triples'), ('HR', 'hr'), ('AB', 'ab')]:
            exp = num(r[at_idx[label]])
            got = career_sum[n][field]
            if exp != got:
                mismatches.append(f'{n} {label}: All-Time={exp} parsed={got}')

    print('=== VALIDATION vs All-Time Stats ===', file=sys.stderr)
    if mismatches:
        for m in mismatches:
            print('  MISMATCH:', m, file=sys.stderr)
    else:
        print('  all 1B/2B/3B/HR/AB career sums match', file=sys.stderr)
    print(f'  players={len(players)} seasons={len(seasons)}', file=sys.stderr)

    json.dump({'players': players, 'seasons': seasons}, sys.stdout, indent=2)


if __name__ == '__main__':
    main()
